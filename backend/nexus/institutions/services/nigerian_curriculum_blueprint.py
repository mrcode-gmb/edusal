"""
Nigerian Tertiary Curriculum Master Blueprint & Hierarchy Importer Service.
Provides standard NUC CCMAS/BMAS, NBTE, and NCCE discipline catalogs,
pre-populated Excel/CSV template generators, and bulk hierarchy provisioning.
"""

from typing import Any, Dict, List, Optional
import io
import csv
from pathlib import Path
from django.db import transaction
from nexus.institutions.models import (
    Institution,
    AcademicDivision,
    Department,
    AcademicProgram,
    AwardLevel,
    TierTwoTerm,
    SiwesPatternChoice,
    SiwesAcademicImpactChoice,
)

ASSETS_DIR = Path(__file__).resolve().parent / "assets"

# ---------------------------------------------------------------------------
# Master Catalog of Standard Nigerian Higher Education Faculties & Programs
# ---------------------------------------------------------------------------

NIGERIAN_MASTER_CATALOG: Dict[str, Dict[str, Any]] = {
    "ENG": {
        "name": "Faculty of Engineering & Technology",
        "code": "ENG",
        "division_type": "FACULTY",
        "archetypes": ["UNIVERSITY_COMPREHENSIVE", "UNIVERSITY_TECHNOLOGY"],
        "departments": [
            {
                "name": "Department of Mechanical Engineering",
                "code": "MEE",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Eng Mechanical Engineering",
                        "code": "BENG-MEE",
                        "award_level": AwardLevel.BENG,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    },
                    {
                        "name": "B.Eng Mechatronics Engineering",
                        "code": "BENG-MTR",
                        "award_level": AwardLevel.BENG,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    },
                ],
            },
            {
                "name": "Department of Civil Engineering",
                "code": "CVE",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Eng Civil Engineering",
                        "code": "BENG-CVE",
                        "award_level": AwardLevel.BENG,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    }
                ],
            },
            {
                "name": "Department of Electrical & Electronics Engineering",
                "code": "EEE",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Eng Electrical & Electronics Engineering",
                        "code": "BENG-EEE",
                        "award_level": AwardLevel.BENG,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    }
                ],
            },
            {
                "name": "Department of Chemical Engineering",
                "code": "CHE",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Eng Chemical Engineering",
                        "code": "BENG-CHE",
                        "award_level": AwardLevel.BENG,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    }
                ],
            },
            {
                "name": "Department of Computer Engineering",
                "code": "CPE",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Eng Computer Engineering",
                        "code": "BENG-CPE",
                        "award_level": AwardLevel.BENG,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    }
                ],
            },
        ],
    },
    "CMP": {
        "name": "Faculty of Computing & Information Technology",
        "code": "CMP",
        "division_type": "FACULTY",
        "archetypes": ["UNIVERSITY_COMPREHENSIVE", "UNIVERSITY_TECHNOLOGY"],
        "departments": [
            {
                "name": "Department of Computer Science",
                "code": "CSC",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc Computer Science",
                        "code": "BSC-CSC",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SEM2_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [3],
                    }
                ],
            },
            {
                "name": "Department of Cybersecurity",
                "code": "CYB",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc Cybersecurity",
                        "code": "BSC-CYB",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SEM2_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [3],
                    }
                ],
            },
            {
                "name": "Department of Software Engineering",
                "code": "SWE",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc Software Engineering",
                        "code": "BSC-SWE",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SEM2_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [3],
                    }
                ],
            },
            {
                "name": "Department of Information Technology",
                "code": "IFT",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc Information Technology",
                        "code": "BSC-IFT",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SEM2_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [3],
                    }
                ],
            },
        ],
    },
    "SCI": {
        "name": "Faculty of Physical & Natural Sciences",
        "code": "SCI",
        "division_type": "FACULTY",
        "archetypes": ["UNIVERSITY_COMPREHENSIVE", "UNIVERSITY_TECHNOLOGY"],
        "departments": [
            {
                "name": "Department of Biochemistry",
                "code": "BCH",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc Biochemistry",
                        "code": "BSC-BCH",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SEM2_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [3],
                    }
                ],
            },
            {
                "name": "Department of Microbiology",
                "code": "MCB",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc Microbiology",
                        "code": "BSC-MCB",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SEM2_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [3],
                    }
                ],
            },
            {
                "name": "Department of Pure & Applied Chemistry",
                "code": "CHM",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc Industrial Chemistry",
                        "code": "BSC-ICH",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SPLIT_200L_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.VACATION_ONLY,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [2, 3],
                    },
                    {
                        "name": "B.Sc Pure Chemistry",
                        "code": "BSC-CHM",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SPLIT_200L_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.VACATION_ONLY,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [2, 3],
                    },
                ],
            },
            {
                "name": "Department of Physics",
                "code": "PHY",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc Physics with Electronics",
                        "code": "BSC-PHE",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SEM2_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [3],
                    }
                ],
            },
        ],
    },
    "ENV": {
        "name": "Faculty of Environmental Sciences",
        "code": "ENV",
        "division_type": "FACULTY",
        "archetypes": ["UNIVERSITY_COMPREHENSIVE", "UNIVERSITY_TECHNOLOGY"],
        "departments": [
            {
                "name": "Department of Architecture",
                "code": "ARC",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc Architecture",
                        "code": "BSC-ARC",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.SPLIT_200L_300L,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.VACATION_ONLY,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [2, 3],
                    }
                ],
            },
            {
                "name": "Department of Quantity Surveying",
                "code": "QSV",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Tech Quantity Surveying",
                        "code": "BTECH-QSV",
                        "award_level": AwardLevel.BTECH,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    }
                ],
            },
            {
                "name": "Department of Urban & Regional Planning",
                "code": "URP",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Tech Urban & Regional Planning",
                        "code": "BTECH-URP",
                        "award_level": AwardLevel.BTECH,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    }
                ],
            },
        ],
    },
    "AGR": {
        "name": "Faculty of Agriculture & Food Security",
        "code": "AGR",
        "division_type": "FACULTY",
        "archetypes": ["UNIVERSITY_COMPREHENSIVE", "UNIVERSITY_TECHNOLOGY"],
        "departments": [
            {
                "name": "Department of Agronomy & Crop Science",
                "code": "CRP",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Agric Agronomy",
                        "code": "BAGR-CRP",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.FULL_SESSION_ATTACHMENT,
                        "siwes_duration_months": 9,
                        "siwes_target_levels": [4],
                    }
                ],
            },
            {
                "name": "Department of Animal Science",
                "code": "ANS",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Agric Animal Science",
                        "code": "BAGR-ANS",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.FULL_SESSION_ATTACHMENT,
                        "siwes_duration_months": 9,
                        "siwes_target_levels": [4],
                    }
                ],
            },
        ],
    },
    "MGT": {
        "name": "Faculty of Management Sciences",
        "code": "MGT",
        "division_type": "FACULTY",
        "archetypes": ["UNIVERSITY_COMPREHENSIVE"],
        "departments": [
            {
                "name": "Department of Business Administration",
                "code": "BUS",
                "siwes_eligible": False,
                "programs": [
                    {
                        "name": "B.Sc Business Administration",
                        "code": "BSC-BUS",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.EXEMPT,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.EXEMPT,
                        "siwes_duration_months": 0,
                        "siwes_target_levels": [],
                    }
                ],
            },
            {
                "name": "Department of Accounting",
                "code": "ACC",
                "siwes_eligible": False,
                "programs": [
                    {
                        "name": "B.Sc Accounting",
                        "code": "BSC-ACC",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.EXEMPT,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.EXEMPT,
                        "siwes_duration_months": 0,
                        "siwes_target_levels": [],
                    }
                ],
            },
        ],
    },
    "BMS": {
        "name": "Faculty of Basic Medical & Allied Health Sciences",
        "code": "BMS",
        "division_type": "FACULTY",
        "archetypes": ["UNIVERSITY_COMPREHENSIVE", "UNIVERSITY_TECHNOLOGY"],
        "departments": [
            {
                "name": "Department of Medical Laboratory Science",
                "code": "MLS",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.MLS Medical Laboratory Science",
                        "code": "BMLS-MLS",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    }
                ],
            },
            {
                "name": "Department of Nursing Science",
                "code": "NUR",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.N.Sc Nursing Science",
                        "code": "BNSC-NUR",
                        "award_level": AwardLevel.BSC,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.YEAR4_400L_EXTENDED,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 6,
                        "siwes_target_levels": [4],
                    }
                ],
            },
        ],
    },
    "LAW": {
        "name": "Faculty of Law",
        "code": "LAW",
        "division_type": "FACULTY",
        "archetypes": ["UNIVERSITY_COMPREHENSIVE"],
        "departments": [
            {
                "name": "Department of Public & Commercial Law",
                "code": "PCL",
                "siwes_eligible": False,
                "programs": [
                    {
                        "name": "LL.B Bachelor of Laws",
                        "code": "LLB-LAW",
                        "award_level": AwardLevel.BA,
                        "duration_years": 5,
                        "siwes_pattern": SiwesPatternChoice.EXEMPT,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.EXEMPT,
                        "siwes_duration_months": 0,
                        "siwes_target_levels": [],
                    }
                ],
            }
        ],
    },
    "EDU": {
        "name": "Faculty of Education",
        "code": "EDU",
        "division_type": "FACULTY",
        "archetypes": ["UNIVERSITY_COMPREHENSIVE", "COLLEGE_OF_EDUCATION"],
        "departments": [
            {
                "name": "Department of Science Education",
                "code": "SED",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Sc (Ed) Computer Science Education",
                        "code": "BED-CSC",
                        "award_level": AwardLevel.BED,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.TEACHING_PRACTICE,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 3,
                        "siwes_target_levels": [3],
                    }
                ],
            },
            {
                "name": "Department of Vocational & Technical Education",
                "code": "VTE",
                "siwes_eligible": True,
                "programs": [
                    {
                        "name": "B.Tech (Ed) Electrical Technology Education",
                        "code": "BED-EED",
                        "award_level": AwardLevel.BED,
                        "duration_years": 4,
                        "siwes_pattern": SiwesPatternChoice.TEACHING_PRACTICE,
                        "siwes_academic_impact": SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE,
                        "siwes_duration_months": 3,
                        "siwes_target_levels": [3],
                    }
                ],
            },
        ],
    },
}

SIWES_PATTERNS_DOCUMENTATION = [
    {
        "code": "SPLIT_200L_300L",
        "title": "Split Vacation (3 Mo @ 200L End + 3 Mo @ 300L End)",
        "levels": "200 Level & 300 Level",
        "total_months": 6,
        "impact": "Vacation Only (0 Semester Disruption)",
        "disciplines": "Pure & Applied Sciences, Industrial Chemistry, Architecture, Applied Geophysics",
        "description": "Students participate in two distinct 3-month attachments during the long vacation after 200L and 300L. No academic semester or lectures are missed.",
    },
    {
        "code": "SEM2_300L",
        "title": "300 Level Second Semester (6 Months Continuous)",
        "levels": "300 Level",
        "total_months": 6,
        "impact": "Replaces Second Semester Coursework",
        "disciplines": "Computer Science, Cybersecurity, Software Engineering, Biochemistry, Microbiology",
        "description": "Standard for 4-year B.Sc. disciplines. Runs continuously throughout the entire Second Semester of 300 Level and extends across the subsequent long vacation.",
    },
    {
        "code": "YEAR4_400L_EXTENDED",
        "title": "400 Level Extended (6 to 9 Months Attachment)",
        "levels": "400 Level",
        "total_months": "6 to 9",
        "impact": "Extended Industrial Year",
        "disciplines": "Mechanical, Civil, Electrical, Chemical Engineering, Agriculture, Medical Lab Science",
        "description": "Standard for 5-year professional degrees (B.Eng / B.Tech / B.Agric / B.MLS). Students spend 6 to 9 continuous months in industry before returning to campus for 500 Level (Final Year).",
    },
    {
        "code": "ND_VACATION",
        "title": "ND Industrial Attachment (3 to 4 Months Vacation)",
        "levels": "ND I / ND II",
        "total_months": "3 to 4",
        "impact": "Vacation Only",
        "disciplines": "Polytechnic National Diploma (ND) Engineering, Sciences & Technology",
        "description": "Runs during the long vacation between ND I and ND II for polytechnic and monotechnic students.",
    },
    {
        "code": "POST_ND_MANDATORY",
        "title": "Post-ND Mandatory Industrial Training (12 Months)",
        "levels": "Post-ND Gap Year",
        "total_months": 12,
        "impact": "1-Year Mandatory Internship",
        "disciplines": "All Polytechnic Graduates seeking Higher National Diploma (HND) admission",
        "description": "Mandatory 1-year industrial training required by NBTE following completion of ND before eligibility for HND admission.",
    },
    {
        "code": "TEACHING_PRACTICE",
        "title": "Teaching Practice / Practicum (3 to 6 Months)",
        "levels": "300L or 400L / NCE II",
        "total_months": "3 to 6",
        "impact": "In-Session Practicum",
        "disciplines": "Faculty of Education (B.Ed / B.Sc.Ed) & Colleges of Education (NCE)",
        "description": "Practical classroom teaching attachment conducted at accredited primary or secondary schools.",
    },
    {
        "code": "EXEMPT",
        "title": "Exempt / Non-Participating (0 Months)",
        "levels": "None",
        "total_months": 0,
        "impact": "No SIWES Requirement",
        "disciplines": "Law (LL.B), Arts & Humanities, Business Administration, Accounting",
        "description": "Academic disciplines that do not participate in the ITF SIWES framework.",
    },
]

DATA_DICTIONARY = [
    ("division_name", "YES", "Name of the School, Faculty, or College", "Text (e.g. Faculty of Engineering)", "Faculty of Engineering"),
    ("division_code", "YES", "Short alphanumeric code for division", "Text (e.g. ENG, CMP, SCI)", "ENG"),
    ("division_type", "YES", "Organizational tier type", "FACULTY, SCHOOL, or COLLEGE", "FACULTY"),
    ("dean_name", "NO", "Full name of the Dean or Provost", "Text (Optional)", "Prof. A. Sani"),
    ("dean_email", "NO", "Official email address of the Dean", "Email format (Optional)", "dean.eng@institution.edu.ng"),
    ("department_name", "YES", "Official title of the academic department", "Text (e.g. Department of Mechanical Engineering)", "Department of Mechanical Engineering"),
    ("department_code", "YES", "Short department code", "Text (e.g. MEE, CSC, CVE)", "MEE"),
    ("hod_name", "NO", "Full name of Head of Department", "Text (Optional)", "Dr. K. Bello"),
    ("hod_email", "NO", "Official email address of the HOD", "Email format (Optional)", "hod.mee@institution.edu.ng"),
    ("siwes_eligible", "YES", "Whether department participates in SIWES", "TRUE or FALSE", "TRUE"),
    ("program_name", "YES", "Full official title of the degree programme", "Text (e.g. B.Eng Mechanical Engineering)", "B.Eng Mechanical Engineering"),
    ("program_code", "YES", "Unique degree programme code", "Text (e.g. BENG-MEE, BSC-CSC)", "BENG-MEE"),
    ("award_level", "YES", "Graduation credential type", "BSC, BENG, BTECH, BA, BED, LLB, MBBS, ND, HND, NCE, PGD, MSC", "BENG"),
    ("duration_years", "YES", "Standard program duration in years", "Integer (1 to 7)", "5"),
    ("siwes_duration_months", "YES", "Total industrial attachment duration in months", "Integer (0 to 24)", "6"),
    ("siwes_pattern", "YES", "Operational placement schedule", "SPLIT_200L_300L, SEM2_300L, YEAR4_400L_EXTENDED, ND_VACATION, POST_ND_MANDATORY, TEACHING_PRACTICE, EXEMPT", "YEAR4_400L_EXTENDED"),
]


# ---------------------------------------------------------------------------
# Blueprint Querying & Provisioning Functions
# ---------------------------------------------------------------------------

def get_master_blueprints(archetype: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Returns the structured master catalog filtered optionally by institutional archetype.
    """
    results = []
    for code, faculty_data in NIGERIAN_MASTER_CATALOG.items():
        if archetype and archetype not in faculty_data["archetypes"]:
            continue

        dept_count = len(faculty_data["departments"])
        prog_count = sum(len(d["programs"]) for d in faculty_data["departments"])
        siwes_dept_count = sum(1 for d in faculty_data["departments"] if d["siwes_eligible"])

        results.append({
            "key": code,
            "name": faculty_data["name"],
            "code": faculty_data["code"],
            "division_type": faculty_data["division_type"],
            "archetypes": faculty_data["archetypes"],
            "departments_count": dept_count,
            "programs_count": prog_count,
            "siwes_departments_count": siwes_dept_count,
            "departments": faculty_data["departments"],
        })
    return results


def import_blueprint_to_institution(
    institution: Institution,
    division_keys: List[str],
) -> Dict[str, int]:
    """
    Idempotently deploys selected blueprint faculties, departments, and degree programmes
    into the specified Institution in an atomic transaction.
    """
    stats = {
        "divisions_created": 0,
        "divisions_reused": 0,
        "departments_created": 0,
        "departments_reused": 0,
        "programs_created": 0,
        "programs_reused": 0,
    }

    import_all = "ALL" in division_keys or not division_keys

    with transaction.atomic():
        for code, faculty_data in NIGERIAN_MASTER_CATALOG.items():
            if not import_all and code not in division_keys:
                continue

            # 1. Get or create Division
            division, div_created = AcademicDivision.objects.get_or_create(
                institution=institution,
                name=faculty_data["name"],
                defaults={
                    "code": faculty_data["code"],
                    "division_type": faculty_data["division_type"],
                    "is_active": True,
                },
            )
            if div_created:
                stats["divisions_created"] += 1
            else:
                stats["divisions_reused"] += 1

            # 2. Get or create Departments
            for dept_data in faculty_data["departments"]:
                department, dept_created = Department.objects.get_or_create(
                    division=division,
                    name=dept_data["name"],
                    defaults={
                        "institution": institution,
                        "code": dept_data["code"],
                        "siwes_eligible": dept_data["siwes_eligible"],
                        "is_active": True,
                    },
                )
                if dept_created:
                    stats["departments_created"] += 1
                else:
                    stats["departments_reused"] += 1

                # 3. Get or create Academic Programmes
                for prog_data in dept_data["programs"]:
                    program, prog_created = AcademicProgram.objects.get_or_create(
                        department=department,
                        name=prog_data["name"],
                        defaults={
                            "institution": institution,
                            "program_code": prog_data["code"],
                            "award_level": prog_data["award_level"],
                            "duration_years": prog_data["duration_years"],
                            "siwes_duration_months": prog_data["siwes_duration_months"],
                            "siwes_pattern": prog_data.get("siwes_pattern", SiwesPatternChoice.SEM2_300L),
                            "siwes_academic_impact": prog_data.get(
                                "siwes_academic_impact", SiwesAcademicImpactChoice.SECOND_SEMESTER_SUBSTITUTE
                            ),
                            "siwes_target_levels": prog_data.get("siwes_target_levels", []),
                            "is_active": True,
                        },
                    )
                    if prog_created:
                        stats["programs_created"] += 1
                    else:
                        stats["programs_reused"] += 1

    return stats


# ---------------------------------------------------------------------------
# Beautiful Multi-Sheet Excel Workbook Generator
# ---------------------------------------------------------------------------

def generate_hierarchy_excel(prepopulate: bool = True, archetype: Optional[str] = None) -> bytes:
    """
    Generates a polished multi-sheet Excel (.xlsx) workbook containing:
    1. INSTRUCTIONS & SIWES PLACEMENT GUIDE (with branding & logo)
    2. ACADEMIC HIERARCHY DATA (pre-populated or structured blank)
    3. REFERENCE LOOKUPS (codes, patterns, award levels)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Color Palette
    BRAND_GREEN = "146B4A"
    BRAND_LIGHT = "F0F7F4"
    ACCENT_MINT = "D8EFE5"
    INK_DARK = "1F2933"
    GRAY_TEXT = "52606D"
    BORDER_COLOR = "DDE3EA"
    WHITE = "FFFFFF"

    font_title = Font(name="Segoe UI", size=16, bold=True, color="146B4A")
    font_subtitle = Font(name="Segoe UI", size=11, bold=True, color="1F2933")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="146B4A")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="1F2933")
    font_regular = Font(name="Segoe UI", size=10, color="1F2933")
    font_muted = Font(name="Segoe UI", size=9, italic=True, color="52606D")

    fill_header = PatternFill(start_color=BRAND_GREEN, end_color=BRAND_GREEN, fill_type="solid")
    fill_zebra = PatternFill(start_color=BRAND_LIGHT, end_color=BRAND_LIGHT, fill_type="solid")
    fill_highlight = PatternFill(start_color=ACCENT_MINT, end_color=ACCENT_MINT, fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    # -----------------------------------------------------------------------
    # SHEET 1: INSTRUCTIONS & SIWES GUIDE
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Instructions & SIWES Guide"
    ws1.views.sheetView[0].showGridLines = True

    # Header & Branding Banner
    ws1.merge_cells("A2:H2")
    ws1["A2"] = "NEXUS EDUTECH PLATFORM"
    ws1["A2"].font = font_title

    ws1.merge_cells("A3:H3")
    ws1["A3"] = "Official 4-Tier Academic Hierarchy & SIWES Placement Onboarding Specification"
    ws1["A3"].font = font_subtitle

    ws1.merge_cells("A4:H4")
    mode_label = "Pre-Populated National Standard" if prepopulate else "Blank Template"
    ws1["A4"] = f"Workbook Mode: {mode_label} · Standard NUC CCMAS/BMAS & NBTE Aligned"
    ws1["A4"].font = font_muted

    # Try attaching company logo image
    logo_path = ASSETS_DIR / "logo.png"
    if logo_path.exists():
        try:
            from openpyxl.drawing.image import Image
            img = Image(str(logo_path))
            img.width = 160
            img.height = 42
            ws1.add_image(img, "I2")
        except Exception:
            pass

    curr_row = 6

    # Section 1: Overview
    ws1.cell(row=curr_row, column=1, value="1. Overview & 4-Tier Hierarchy Architecture").font = font_section
    curr_row += 1
    overview_text = (
        "This workbook enables institutional administrators to rapidly define and configure their entire academic structure "
        "without manual single-item entry. The platform organizes higher education governance into 4 strict relational tiers:\n"
        " • Tier 1 (Institution): The root university, polytechnic, or college.\n"
        " • Tier 2 (Division / Faculty): Schools, Faculties, or Colleges (e.g. Faculty of Engineering).\n"
        " • Tier 3 (Department): Academic and SIWES-eligible departments (e.g. Department of Computer Science).\n"
        " • Tier 4 (Degree Programme): Specific degree options, award levels, durations, and SIWES operational placement patterns."
    )
    ws1.cell(row=curr_row, column=1, value=overview_text).font = font_regular
    curr_row += 2

    # Section 2: SIWES Placement Guide Table
    ws1.cell(row=curr_row, column=1, value="2. SIWES Operational Placement Patterns Guide (All Available Options)").font = font_section
    curr_row += 1

    siwes_headers = ["Pattern Code", "Official Title", "Placement Level(s)", "Total Duration", "Calendar Impact", "Applicable Disciplines", "Operational Description"]
    for c_idx, h in enumerate(siwes_headers, start=1):
        cell = ws1.cell(row=curr_row, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws1.row_dimensions[curr_row].height = 24
    curr_row += 1

    for item in SIWES_PATTERNS_DOCUMENTATION:
        row_vals = [
            item["code"],
            item["title"],
            item["levels"],
            f"{item['total_months']} Months" if isinstance(item["total_months"], int) and item["total_months"] > 0 else f"{item['total_months']} Months" if str(item["total_months"]).isdigit() else item["total_months"],
            item["impact"],
            item["disciplines"],
            item["description"],
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=curr_row, column=c_idx, value=str(val))
            cell.font = font_bold if c_idx == 1 else font_regular
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws1.row_dimensions[curr_row].height = 42
        curr_row += 1

    curr_row += 2

    # Section 3: Data Dictionary
    ws1.cell(row=curr_row, column=1, value="3. Template Column Specifications & Data Dictionary").font = font_section
    curr_row += 1

    dict_headers = ["Column Header", "Required?", "Field Description", "Allowed Values / Options", "Example Value"]
    for c_idx, h in enumerate(dict_headers, start=1):
        cell = ws1.cell(row=curr_row, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws1.row_dimensions[curr_row].height = 22
    curr_row += 1

    for row_tuple in DATA_DICTIONARY:
        for c_idx, val in enumerate(row_tuple, start=1):
            cell = ws1.cell(row=curr_row, column=c_idx, value=val)
            cell.font = font_bold if c_idx in [1, 2] else font_regular
            if c_idx == 2 and val == "YES":
                cell.fill = fill_highlight
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
        curr_row += 1

    curr_row += 2

    # Section 4: Quick Workflow
    ws1.cell(row=curr_row, column=1, value="4. How to Populate & Import This Template").font = font_section
    curr_row += 1
    instructions = [
        "1. Switch to the 'Academic Hierarchy Data' sheet in this workbook.",
        "2. Pre-Populated Template: All standard Nigerian university disciplines are already filled in with correct NUC codes, award levels, durations, and SIWES patterns. Simply remove faculties or programmes you do NOT offer, and fill in specific Dean/HOD names.",
        "3. Blank Template: Fill in each row with your institution's specific faculties, departments, and programmes.",
        "4. Save this file as .xlsx or .csv on your computer.",
        "5. Go to Nexus Portal -> Hierarchy Explorer -> Click 'Bulk Setup Wizard (Excel / Blueprint)' -> Upload file -> Confirm import.",
    ]
    for step in instructions:
        ws1.cell(row=curr_row, column=1, value=step).font = font_regular
        curr_row += 1

    # Auto-fit column widths for Sheet 1
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 38
    ws1.column_dimensions["C"].width = 24
    ws1.column_dimensions["D"].width = 20
    ws1.column_dimensions["E"].width = 32
    ws1.column_dimensions["F"].width = 40
    ws1.column_dimensions["G"].width = 50

    # -----------------------------------------------------------------------
    # SHEET 2: ACADEMIC HIERARCHY DATA
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Academic Hierarchy Data")
    ws2.views.sheetView[0].showGridLines = True

    headers = [
        "division_name",
        "division_code",
        "division_type",
        "dean_name",
        "dean_email",
        "department_name",
        "department_code",
        "hod_name",
        "hod_email",
        "siwes_eligible",
        "program_name",
        "program_code",
        "award_level",
        "duration_years",
        "siwes_duration_months",
        "siwes_pattern",
    ]

    # Write Header Row
    for col_idx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws2.row_dimensions[1].height = 28

    data_row = 2
    if prepopulate:
        for code, faculty_data in NIGERIAN_MASTER_CATALOG.items():
            if archetype and archetype not in faculty_data["archetypes"]:
                continue
            for dept in faculty_data["departments"]:
                for prog in dept["programs"]:
                    row_values = [
                        faculty_data["name"],
                        faculty_data["code"],
                        faculty_data["division_type"],
                        "",  # dean_name
                        "",  # dean_email
                        dept["name"],
                        dept["code"],
                        "",  # hod_name
                        "",  # hod_email
                        "TRUE" if dept["siwes_eligible"] else "FALSE",
                        prog["name"],
                        prog["code"],
                        prog["award_level"],
                        prog["duration_years"],
                        prog["siwes_duration_months"],
                        prog.get("siwes_pattern", SiwesPatternChoice.SEM2_300L),
                    ]
                    for c_idx, val in enumerate(row_values, start=1):
                        cell = ws2.cell(row=data_row, column=c_idx, value=val)
                        cell.font = font_regular
                        cell.border = thin_border
                        if data_row % 2 == 0:
                            cell.fill = fill_zebra
                        if c_idx in [2, 3, 7, 10, 12, 13, 14, 15, 16]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws2.row_dimensions[data_row].height = 20
                    data_row += 1
    else:
        # Sample Guide Rows for Blank Mode
        sample_rows = [
            ("Faculty of Engineering", "ENG", "FACULTY", "Prof. A. Sani", "dean.eng@uni.edu", "Department of Mechanical Engineering", "MEE", "Dr. K. Bello", "hod.mee@uni.edu", "TRUE", "B.Eng Mechanical Engineering", "BENG-MEE", "BENG", 5, 6, "YEAR4_400L_EXTENDED"),
            ("Faculty of Computing", "CMP", "FACULTY", "Prof. C. Eze", "dean.cmp@uni.edu", "Department of Computer Science", "CSC", "Dr. T. Alabi", "hod.csc@uni.edu", "TRUE", "B.Sc Computer Science", "BSC-CSC", "BSC", 4, 6, "SEM2_300L"),
            ("Faculty of Science", "SCI", "FACULTY", "Prof. H. Yusuf", "dean.sci@uni.edu", "Department of Industrial Chemistry", "ICH", "Dr. R. Adamu", "hod.ich@uni.edu", "TRUE", "B.Sc Industrial Chemistry", "BSC-ICH", "BSC", 4, 6, "SPLIT_200L_300L"),
        ]
        for s_idx, s_row in enumerate(sample_rows, start=2):
            for c_idx, val in enumerate(s_row, start=1):
                cell = ws2.cell(row=s_idx, column=c_idx, value=val)
                cell.font = font_regular
                cell.border = thin_border
                cell.fill = fill_highlight
            ws2.row_dimensions[s_idx].height = 20
            data_row += 1

    # Auto-adjust column widths
    column_widths = {
        "A": 36,  # division_name
        "B": 16,  # division_code
        "C": 16,  # division_type
        "D": 22,  # dean_name
        "E": 26,  # dean_email
        "F": 42,  # department_name
        "G": 18,  # department_code
        "H": 22,  # hod_name
        "I": 26,  # hod_email
        "J": 16,  # siwes_eligible
        "K": 38,  # program_name
        "L": 18,  # program_code
        "M": 15,  # award_level
        "N": 16,  # duration_years
        "O": 22,  # siwes_duration_months
        "P": 28,  # siwes_pattern
    }
    for col_letter, width in column_widths.items():
        ws2.column_dimensions[col_letter].width = width

    # -----------------------------------------------------------------------
    # SHEET 3: REFERENCE LOOKUPS
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet(title="Lookup Codes & Choices")
    ws3.views.sheetView[0].showGridLines = True

    ws3.cell(row=1, column=1, value="Valid Lookup Values & Taxonomy Reference").font = font_section

    # Award Levels
    ws3.cell(row=3, column=1, value="Award Level Codes").font = font_bold
    ws3.cell(row=3, column=2, value="Full Award Description").font = font_bold
    ws3.cell(row=3, column=1).fill = fill_highlight
    ws3.cell(row=3, column=2).fill = fill_highlight
    award_items = [
        ("BSC", "Bachelor of Science (B.Sc.)"),
        ("BENG", "Bachelor of Engineering (B.Eng.)"),
        ("BTECH", "Bachelor of Technology (B.Tech.)"),
        ("BA", "Bachelor of Arts (B.A.)"),
        ("BED", "Bachelor of Education (B.Ed.)"),
        ("LLB", "Bachelor of Laws (LL.B.)"),
        ("MBBS", "Bachelor of Medicine & Surgery (MBBS)"),
        ("ND", "National Diploma (ND)"),
        ("HND", "Higher National Diploma (HND)"),
        ("NCE", "Nigeria Certificate in Education (NCE)"),
        ("PGD", "Postgraduate Diploma (PGD)"),
        ("MSC", "Master of Science (M.Sc.)"),
    ]
    for idx, (k, v) in enumerate(award_items, start=4):
        ws3.cell(row=idx, column=1, value=k).font = font_bold
        ws3.cell(row=idx, column=2, value=v).font = font_regular
        ws3.cell(row=idx, column=1).border = thin_border
        ws3.cell(row=idx, column=2).border = thin_border

    # SIWES Patterns
    start_r = len(award_items) + 6
    ws3.cell(row=start_r, column=1, value="SIWES Pattern Codes").font = font_bold
    ws3.cell(row=start_r, column=2, value="Full SIWES Operational Schedule").font = font_bold
    ws3.cell(row=start_r, column=1).fill = fill_highlight
    ws3.cell(row=start_r, column=2).fill = fill_highlight
    for idx, item in enumerate(SIWES_PATTERNS_DOCUMENTATION, start=start_r + 1):
        ws3.cell(row=idx, column=1, value=item["code"]).font = font_bold
        ws3.cell(row=idx, column=2, value=item["title"]).font = font_regular
        ws3.cell(row=idx, column=1).border = thin_border
        ws3.cell(row=idx, column=2).border = thin_border

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 55

    # Export workbook to bytes buffer
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def generate_hierarchy_csv(prepopulate: bool = True, archetype: Optional[str] = None) -> str:
    """
    Generates a pre-populated or empty CSV spreadsheet template with standard columns.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    headers = [
        "division_name",
        "division_code",
        "division_type",
        "dean_name",
        "dean_email",
        "department_name",
        "department_code",
        "hod_name",
        "hod_email",
        "siwes_eligible",
        "program_name",
        "program_code",
        "award_level",
        "duration_years",
        "siwes_duration_months",
        "siwes_pattern",
    ]
    writer.writerow(headers)

    if prepopulate:
        for code, faculty_data in NIGERIAN_MASTER_CATALOG.items():
            if archetype and archetype not in faculty_data["archetypes"]:
                continue
            for dept in faculty_data["departments"]:
                for prog in dept["programs"]:
                    writer.writerow([
                        faculty_data["name"],
                        faculty_data["code"],
                        faculty_data["division_type"],
                        "",  # dean_name left for admin to customize
                        "",  # dean_email
                        dept["name"],
                        dept["code"],
                        "",  # hod_name left for admin to customize
                        "",  # hod_email
                        "TRUE" if dept["siwes_eligible"] else "FALSE",
                        prog["name"],
                        prog["code"],
                        prog["award_level"],
                        prog["duration_years"],
                        prog["siwes_duration_months"],
                        prog.get("siwes_pattern", SiwesPatternChoice.SEM2_300L),
                    ])

    return output.getvalue()
