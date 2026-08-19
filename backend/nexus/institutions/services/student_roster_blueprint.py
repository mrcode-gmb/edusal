"""
Service for generating program-specific student onboarding Excel/CSV templates
and parsing/validating bulk cohort uploads.
"""

from __future__ import annotations
import io
import os
import csv
import re
from typing import Any, Dict, List, Optional, Tuple
from django.db import transaction
from django.utils.text import slugify

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

from nexus.institutions.models import (
    Institution,
    AcademicDivision,
    Department,
    AcademicProgram,
    AcademicSession,
    StudentProfile,
    EntryMode,
)
from nexus.users.models import User


# Visual Palette — Emerald & Institutional Slate
COLOR_PRIMARY_DARK = "0D472B"
COLOR_PRIMARY_MID = "146B4A"
COLOR_PRIMARY_LIGHT = "E6F4EA"
COLOR_ACCENT = "D4A017"
COLOR_TEXT_WHITE = "FFFFFF"
COLOR_TEXT_DARK = "1E293B"
COLOR_MUTED = "64748B"
COLOR_BORDER = "CBD5E1"
COLOR_ROW_ALT = "F8FAFC"
COLOR_NOTE_BG = "FEF9C3"
COLOR_NOTE_TEXT = "713F12"


def _make_border(color: str = COLOR_BORDER) -> Border:
    thin = Side(border_style="thin", color=color)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_program_student_excel(program: AcademicProgram) -> bytes:
    """
    Generates a beautifully styled, 3-worksheet Master Student Onboarding Workbook
    tailored specifically to a single Academic Degree Programme.
    """
    wb = openpyxl.Workbook()

    # -------------------------------------------------------------------------
    # Sheet 1: Instructions & Program Info
    # -------------------------------------------------------------------------
    ws_info = wb.active
    ws_info.title = "📘 Instructions & Info"
    ws_info.views.sheetView[0].showGridLines = True

    # 1. Header with Branding
    ws_info.row_dimensions[1].height = 20
    ws_info.row_dimensions[2].height = 45
    ws_info.merge_cells("A2:H2")
    cell_title = ws_info["A2"]
    cell_title.value = f"NEXUS EDUSAL — {program.institution.name.upper()}"
    cell_title.font = Font(name="Arial", size=15, bold=True, color=COLOR_TEXT_WHITE)
    cell_title.fill = PatternFill("solid", fgColor=COLOR_PRIMARY_DARK)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    ws_info.row_dimensions[3].height = 28
    ws_info.merge_cells("A3:H3")
    cell_sub = ws_info["A3"]
    cell_sub.value = (
        f"Official Student Cohort Onboarding Template — {program.name} ({program.program_code})"
    )
    cell_sub.font = Font(name="Arial", size=11, bold=True, color=COLOR_TEXT_WHITE)
    cell_sub.fill = PatternFill("solid", fgColor=COLOR_PRIMARY_MID)
    cell_sub.alignment = Alignment(horizontal="center", vertical="center")

    # Embed Logo if available
    logo_path = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
    if os.path.exists(logo_path):
        try:
            img = OpenpyxlImage(logo_path)
            img.width = 160
            img.height = 40
            ws_info.add_image(img, "H2")
        except Exception:
            pass

    # 2. Programme Metadata Summary Box
    ws_info["A5"] = "🏛️ TARGET PROGRAMME SPECIFICATIONS (SYSTEM VERIFIED)"
    ws_info["A5"].font = Font(name="Arial", size=12, bold=True, color=COLOR_PRIMARY_DARK)

    specs = [
        ("Institution Name:", program.institution.name),
        ("Academic Division (Faculty / School):", program.department.division.name),
        ("Department:", program.department.name),
        ("Degree Programme Name:", program.name),
        ("Programme Code:", program.program_code or "N/A"),
        ("Award Level:", program.get_award_level_display() if hasattr(program, "get_award_level_display") else program.award_level),
        ("Duration of Study:", f"{program.duration_years} Academic Years"),
        ("SIWES Training Duration:", f"{program.siwes_duration_months} Months ({program.get_siwes_pattern_display() if hasattr(program, 'get_siwes_pattern_display') else 'Standard'})"),
        ("System Program UUID (DO NOT EDIT):", str(program.id)),
    ]

    for idx, (label, val) in enumerate(specs, start=6):
        ws_info.row_dimensions[idx].height = 22
        ws_info[f"A{idx}"] = label
        ws_info[f"A{idx}"].font = Font(name="Arial", size=10, bold=True, color=COLOR_TEXT_DARK)
        ws_info[f"A{idx}"].fill = PatternFill("solid", fgColor=COLOR_PRIMARY_LIGHT)
        ws_info[f"A{idx}"].border = _make_border()

        ws_info.merge_cells(f"B{idx}:H{idx}")
        cell_v = ws_info[f"B{idx}"]
        cell_v.value = val
        cell_v.font = Font(name="Arial", size=10, bold=(label.startswith("System")), color=COLOR_PRIMARY_DARK if label.startswith("System") else COLOR_TEXT_DARK)
        cell_v.border = _make_border()

    # 3. Step-by-Step Instructions
    row_inst_start = len(specs) + 8
    ws_info[f"A{row_inst_start}"] = "📝 INSTRUCTIONS FOR DEPARTMENTAL OFFICERS & ADMISSIONS STAFF"
    ws_info[f"A{row_inst_start}"].font = Font(name="Arial", size=12, bold=True, color=COLOR_PRIMARY_DARK)

    instructions = [
        "1. Open the '🎓 Student Roster' sheet to enter or paste student records for this degree programme.",
        "2. MATRIC NUMBER: Must be the official university matriculation number (e.g. 2023/SWE/001). Must be unique.",
        "3. FULL NAME: Enter student's official full name (e.g. Adebayo Chukwuma Musa or Musa, Adebayo C.).",
        "4. STUDENT EMAIL: Personal emails (e.g. @gmail.com, @yahoo.com) or official institutional emails are both accepted.",
        f"5. CURRENT LEVEL: Choose between 100L up to {program.duration_years * 100}L (or ND1/ND2/HND1/HND2).",
        "6. ENTRY SESSION: Enter the session of admission (e.g. 2024/2025, 2023/2024). Refer to '📋 Lookup Choices' sheet for active sessions.",
        "7. ENTRY MODE: Choose from UTME, DIRECT_ENTRY, TRANSFER, POSTGRADUATE, or PART_TIME.",
        "8. OPTIONAL FIELDS: Phone Number, CGPA (0.00-5.00), JAMB Reg Number, Gender, and State of Origin can be included if available.",
        "9. SAVING & UPLOADING: Save this workbook and upload it directly on the Nexus Portal under 'Student Roster' -> 'Bulk Import Cohort'.",
    ]

    for idx, inst in enumerate(instructions, start=row_inst_start + 1):
        ws_info.row_dimensions[idx].height = 22
        ws_info.merge_cells(f"A{idx}:H{idx}")
        cell = ws_info[f"A{idx}"]
        cell.value = inst
        cell.font = Font(name="Arial", size=10, color=COLOR_TEXT_DARK)
        cell.alignment = Alignment(vertical="center")

    ws_info.column_dimensions["A"].width = 38
    for col_l in ["B", "C", "D", "E", "F", "G", "H"]:
        ws_info.column_dimensions[col_l].width = 24

    # -------------------------------------------------------------------------
    # Sheet 2: Student Roster
    # -------------------------------------------------------------------------
    ws_data = wb.create_sheet(title="🎓 Student Roster")
    ws_data.views.sheetView[0].showGridLines = True

    # Header Row 1: Program Banner
    ws_data.row_dimensions[1].height = 30
    ws_data.merge_cells("A1:K1")
    ws_data["A1"] = f"STUDENT COHORT ROSTER — {program.name.upper()} ({program.program_code or 'PROG'})"
    ws_data["A1"].font = Font(name="Arial", size=11, bold=True, color=COLOR_TEXT_WHITE)
    ws_data["A1"].fill = PatternFill("solid", fgColor=COLOR_PRIMARY_DARK)
    ws_data["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Header Row 2: Columns
    ws_data.row_dimensions[2].height = 32
    columns = [
        ("matric_number", "Matric Number *", 22),
        ("full_name", "Full Name *", 28),
        ("student_email", "Student Email *", 30),
        ("phone_number", "Phone Number", 18),
        ("current_level", "Current Level *", 16),
        ("entry_session", "Entry Session *", 18),
        ("entry_mode", "Entry Mode *", 18),
        ("cgpa", "CGPA (0.00-5.00)", 18),
        ("jamb_reg_number", "JAMB Reg Number", 22),
        ("gender", "Gender", 14),
        ("state_of_origin", "State of Origin", 18),
    ]

    for col_idx, (col_key, col_title, col_width) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws_data.column_dimensions[col_letter].width = col_width

        cell = ws_data.cell(row=2, column=col_idx)
        cell.value = col_title
        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_TEXT_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _make_border()

    # Active Sessions lookup for sample rows
    sessions = list(AcademicSession.objects.filter(institution=program.institution).order_by("-is_current", "-start_date"))
    sample_session = sessions[0].session_label if sessions else "2024/2025"

    sample_rows = [
        [
            f"2023/{program.program_code or 'STU'}/001",
            "Adebayo Chukwuma Musa",
            "adebayo.musa@gmail.com",
            "08031234567",
            "200L",
            sample_session,
            "UTME",
            "4.25",
            "202310892019AB",
            "MALE",
            "Lagos",
        ],
        [
            f"2023/{program.program_code or 'STU'}/002",
            "Fatima Zainab Bello",
            "fatima.bello@yahoo.com",
            "08099887766",
            "200L",
            sample_session,
            "DIRECT_ENTRY",
            "4.80",
            "202310481920EF",
            "FEMALE",
            "Kano",
        ],
    ]

    for row_idx, rdata in enumerate(sample_rows, start=3):
        ws_data.row_dimensions[row_idx].height = 22
        for col_idx, val in enumerate(rdata, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = Font(name="Arial", size=10, color=COLOR_TEXT_DARK)
            cell.alignment = Alignment(vertical="center", horizontal="left")
            cell.fill = PatternFill("solid", fgColor=COLOR_ROW_ALT)
            cell.border = _make_border()

    # Empty Rows for User Entry (Rows 5 to 50 with borders)
    for row_idx in range(5, 51):
        ws_data.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(columns) + 1):
            cell = ws_data.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10, color=COLOR_TEXT_DARK)
            cell.alignment = Alignment(vertical="center", horizontal="left")
            cell.border = _make_border()

    # -------------------------------------------------------------------------
    # Sheet 3: Lookup Choices & Metadata
    # -------------------------------------------------------------------------
    ws_lookup = wb.create_sheet(title="📋 Lookup Choices")
    ws_lookup.views.sheetView[0].showGridLines = True

    # 1. Level Choices
    ws_lookup["A1"] = "Current Level Choices"
    ws_lookup["A1"].font = Font(name="Arial", size=11, bold=True, color=COLOR_TEXT_WHITE)
    ws_lookup["A1"].fill = PatternFill("solid", fgColor=COLOR_PRIMARY_MID)
    ws_lookup["A1"].alignment = Alignment(horizontal="center", vertical="center")

    level_choices = [f"{y * 100}L" for y in range(1, program.duration_years + 1)]
    if program.award_level in ["ND", "HND"]:
        level_choices = ["ND1", "ND2"] if program.award_level == "ND" else ["HND1", "HND2"]

    for idx, lvl in enumerate(level_choices, start=2):
        ws_lookup[f"A{idx}"] = lvl
        ws_lookup[f"A{idx}"].font = Font(name="Arial", size=10)
        ws_lookup[f"A{idx}"].border = _make_border()

    # 2. Entry Modes
    ws_lookup["C1"] = "Entry Mode Code"
    ws_lookup["D1"] = "Entry Mode Label"
    for col in ["C1", "D1"]:
        ws_lookup[col].font = Font(name="Arial", size=11, bold=True, color=COLOR_TEXT_WHITE)
        ws_lookup[col].fill = PatternFill("solid", fgColor=COLOR_PRIMARY_MID)
        ws_lookup[col].alignment = Alignment(horizontal="center", vertical="center")

    entry_modes = [
        ("UTME", "Unified Tertiary Matriculation Exam (100L Entry)"),
        ("DIRECT_ENTRY", "Direct Entry (200L / Higher Entry)"),
        ("TRANSFER", "Inter-University / Inter-Departmental Transfer"),
        ("POSTGRADUATE", "Postgraduate / Master / Ph.D Entry"),
        ("PART_TIME", "Part-Time / Distance Learning Entry"),
    ]
    for idx, (code, label) in enumerate(entry_modes, start=2):
        ws_lookup[f"C{idx}"] = code
        ws_lookup[f"D{idx}"] = label
        ws_lookup[f"C{idx}"].border = _make_border()
        ws_lookup[f"D{idx}"].border = _make_border()

    # 3. Available Academic Sessions
    ws_lookup["F1"] = "Session Label"
    ws_lookup["G1"] = "Current Active Session?"
    for col in ["F1", "G1"]:
        ws_lookup[col].font = Font(name="Arial", size=11, bold=True, color=COLOR_TEXT_WHITE)
        ws_lookup[col].fill = PatternFill("solid", fgColor=COLOR_PRIMARY_MID)
        ws_lookup[col].alignment = Alignment(horizontal="center", vertical="center")

    for idx, sess in enumerate(sessions, start=2):
        ws_lookup[f"F{idx}"] = sess.session_label
        ws_lookup[f"G{idx}"] = "YES (Current)" if sess.is_current else "NO"
        ws_lookup[f"F{idx}"].border = _make_border()
        ws_lookup[f"G{idx}"].border = _make_border()

    # 4. Hidden System Metadata Block
    ws_lookup["I1"] = "__SYSTEM_METADATA_KEY__"
    ws_lookup["J1"] = "__SYSTEM_METADATA_VAL__"
    ws_lookup["I2"] = "PROGRAM_UUID"
    ws_lookup["J2"] = str(program.id)
    ws_lookup["I3"] = "PROGRAM_CODE"
    ws_lookup["J3"] = str(program.program_code or "")
    ws_lookup["I4"] = "INSTITUTION_ID"
    ws_lookup["J4"] = str(program.institution_id)

    ws_lookup.column_dimensions["A"].width = 24
    ws_lookup.column_dimensions["C"].width = 20
    ws_lookup.column_dimensions["D"].width = 45
    ws_lookup.column_dimensions["F"].width = 20
    ws_lookup.column_dimensions["G"].width = 24
    ws_lookup.column_dimensions["I"].width = 26
    ws_lookup.column_dimensions["J"].width = 40

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_program_student_csv(program: AcademicProgram) -> str:
    """
    Generates a clean CSV template pre-tagged with the specific program code.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header metadata row
    writer.writerow([f"# PROGRAM_UUID: {program.id} | PROGRAM: {program.name} ({program.program_code}) | DEPT: {program.department.name}"])
    
    headers = [
        "matric_number",
        "full_name",
        "student_email",
        "phone_number",
        "current_level",
        "entry_session",
        "entry_mode",
        "cgpa",
        "jamb_reg_number",
        "gender",
        "state_of_origin",
    ]
    writer.writerow(headers)

    sessions = list(AcademicSession.objects.filter(institution=program.institution).order_by("-is_current"))
    sample_sess = sessions[0].session_label if sessions else "2024/2025"

    writer.writerow([
        f"2023/{program.program_code or 'STU'}/001",
        "Adebayo Chukwuma Musa",
        "adebayo.musa@gmail.com",
        "08031234567",
        "200L",
        sample_sess,
        "UTME",
        "4.25",
        "202310892019AB",
        "MALE",
        "Lagos",
    ])
    writer.writerow([
        f"2023/{program.program_code or 'STU'}/002",
        "Fatima Zainab Bello",
        "fatima.bello@yahoo.com",
        "08099887766",
        "200L",
        sample_sess,
        "DIRECT_ENTRY",
        "4.80",
        "202310481920EF",
        "FEMALE",
        "Kano",
    ])

    return output.getvalue()


def parse_and_validate_student_roster(
    file_obj: Any,
    filename: str,
    institution: Institution,
    fallback_program_id: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Parses an uploaded Excel (.xlsx) or CSV student cohort file.
    Extracts embedded program metadata, validates each student record against
    academic constraints, and checks for duplicates.
    """
    raw_rows: List[Dict[str, Any]] = []
    detected_program_id: Optional[str] = None
    fn = filename.lower()

    if fn.endswith(".xlsx") or fn.endswith(".xls"):
        wb = openpyxl.load_workbook(file_obj, data_only=True)

        # 1. Check for Embedded Metadata in Lookup Sheet
        if "📋 Lookup Choices" in wb.sheetnames:
            ws_meta = wb["📋 Lookup Choices"]
            for row in ws_meta.iter_rows(values_only=True):
                if len(row) >= 10:
                    k, v = row[8], row[9]
                    if str(k or "").strip() == "PROGRAM_UUID" and v:
                        detected_program_id = str(v).strip()

        # 2. Extract Data from Student Roster Sheet
        target_sheet = "🎓 Student Roster" if "🎓 Student Roster" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[target_sheet]

        rows_iter = list(ws.iter_rows(values_only=True))
        if len(rows_iter) > 1:
            header_row_idx = 0
            for idx, r in enumerate(rows_iter):
                r_strs = [str(c or "").lower() for c in r]
                if any("matric" in c for c in r_strs) and any("name" in c for c in r_strs):
                    header_row_idx = idx
                    break

            headers = [slugify(str(h or "")).replace("-", "_") for h in rows_iter[header_row_idx]]

            for r in rows_iter[header_row_idx + 1:]:
                if not any(r):
                    continue
                row_dict: Dict[str, Any] = {}
                for idx, h in enumerate(headers):
                    if idx < len(r) and h:
                        val = r[idx]
                        row_dict[h] = str(val).strip() if val is not None else ""

                if row_dict.get("matric_number") or row_dict.get("full_name") or row_dict.get("student_email"):
                    raw_rows.append(row_dict)

    else:
        content = file_obj.read().decode("utf-8-sig", errors="ignore")
        lines = content.splitlines()
        csv_lines = []
        for line in lines:
            if line.strip().startswith("#"):
                match = re.search(r"PROGRAM_UUID:\s*([a-f0-9\-]+)", line, re.IGNORECASE)
                if match:
                    detected_program_id = match.group(1).strip()
            else:
                csv_lines.append(line)

        reader = csv.DictReader(io.StringIO("\n".join(csv_lines)))
        for r in reader:
            normalized = {slugify(k or "").replace("-", "_"): (v or "").strip() for k, v in r.items() if k}
            if any(normalized.values()):
                raw_rows.append(normalized)

    target_program_id = detected_program_id or fallback_program_id

    # 3. Resolve Academic Programme
    program = None
    if target_program_id:
        try:
            program = AcademicProgram.objects.select_related("department", "department__division").get(
                id=target_program_id, institution=institution
            )
        except AcademicProgram.DoesNotExist:
            program = None

    if not program:
        return {
            "success": False,
            "error": "Could not identify the target degree programme. Please select the specific programme before uploading.",
            "detected_program": None,
            "valid_rows": [],
            "errors": [{"row": 0, "message": "Degree programme could not be resolved from file metadata or selection."}],
            "stats": {"total_rows": len(raw_rows), "valid_count": 0, "error_count": len(raw_rows)},
        }

    # 4. Academic Sessions Cache
    sessions_map: Dict[str, AcademicSession] = {}
    for s in AcademicSession.objects.filter(institution=institution):
        clean_lbl = s.session_label.strip().replace(" ", "")
        sessions_map[clean_lbl] = s
        sessions_map[clean_lbl.replace("/", "-")] = s
        sessions_map[clean_lbl.replace("-", "/")] = s

    default_session = AcademicSession.objects.filter(institution=institution, is_current=True).first()
    if not default_session:
        default_session = AcademicSession.objects.filter(institution=institution).first()

    # 5. Row-by-Row Validation
    valid_rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen_matrics: set = set()
    seen_emails: set = set()

    for idx, r in enumerate(raw_rows, start=1):
        matric = (r.get("matric_number") or "").strip().upper()
        name = (r.get("full_name") or r.get("name") or "").strip()
        email = (r.get("student_email") or r.get("email") or "").strip().lower()
        phone = (r.get("phone_number") or r.get("phone") or "").strip()
        level_raw = (r.get("current_level") or r.get("level") or "100L").strip().upper()
        sess_raw = (r.get("entry_session") or r.get("session") or "").strip()
        mode_raw = (r.get("entry_mode") or "UTME").strip().upper().replace(" ", "_")
        cgpa_raw = (r.get("cgpa") or "").strip()
        jamb_reg = (r.get("jamb_reg_number") or r.get("jamb_no") or "").strip().upper()
        gender_raw = (r.get("gender") or "").strip().upper()
        state_origin = (r.get("state_of_origin") or r.get("state") or "").strip()

        row_errors: List[str] = []

        if not matric:
            row_errors.append("Matric number is required.")
        elif matric in seen_matrics:
            row_errors.append(f"Duplicate matric number '{matric}' in this file.")
        seen_matrics.add(matric)

        if not name:
            row_errors.append("Full name is required.")

        if not email or "@" not in email or "." not in email:
            row_errors.append("Valid student email is required.")
        elif email in seen_emails:
            row_errors.append(f"Duplicate email '{email}' in this file.")
        seen_emails.add(email)

        # Parse Level (e.g. 100L -> 1, 200L -> 2, ND1 -> 1, HND1 -> 3)
        year_of_study = 1
        level_match = re.search(r"(\d)00", level_raw)
        if level_match:
            year_of_study = int(level_match.group(1))
        elif "ND1" in level_raw or "YEAR 1" in level_raw:
            year_of_study = 1
        elif "ND2" in level_raw or "YEAR 2" in level_raw:
            year_of_study = 2
        elif "HND1" in level_raw:
            year_of_study = 3
        elif "HND2" in level_raw:
            year_of_study = 4
        else:
            try:
                year_of_study = int(re.sub(r"\D", "", level_raw) or 1)
            except ValueError:
                year_of_study = 1

        if year_of_study > program.duration_years and program.award_level not in ["HND"]:
            row_errors.append(
                f"Level '{level_raw}' exceeds maximum programme duration ({program.duration_years} Years)."
            )

        # Resolve Academic Session
        resolved_session = default_session
        if sess_raw:
            clean_s = sess_raw.replace(" ", "")
            if clean_s in sessions_map:
                resolved_session = sessions_map[clean_s]
            elif default_session:
                resolved_session = default_session
            else:
                row_errors.append(f"Academic session '{sess_raw}' not recognized.")

        # Parse CGPA
        parsed_cgpa = None
        if cgpa_raw:
            try:
                c_val = float(cgpa_raw)
                if 0.0 <= c_val <= 5.0:
                    parsed_cgpa = c_val
                else:
                    row_errors.append(f"CGPA '{cgpa_raw}' must be between 0.00 and 5.00.")
            except ValueError:
                row_errors.append(f"Invalid CGPA format '{cgpa_raw}'.")

        # Entry Mode
        entry_mode_val = EntryMode.UTME
        if mode_raw in [e.value for e in EntryMode]:
            entry_mode_val = mode_raw
        elif "DIRECT" in mode_raw:
            entry_mode_val = EntryMode.DIRECT_ENTRY
        elif "TRANSFER" in mode_raw:
            entry_mode_val = EntryMode.TRANSFER

        # Gender
        gender_val = "MALE" if gender_raw.startswith("M") else "FEMALE" if gender_raw.startswith("F") else "OTHER"

        if row_errors:
            errors.append({
                "row_number": idx,
                "matric_number": matric or "N/A",
                "name": name or "N/A",
                "email": email or "N/A",
                "reasons": row_errors,
            })
        else:
            valid_rows.append({
                "matric_number": matric,
                "name": name,
                "email": email,
                "phone_number": phone,
                "year_of_study": year_of_study,
                "current_level_label": f"{year_of_study * 100}L",
                "entry_session_id": resolved_session.id if resolved_session else None,
                "entry_session_label": resolved_session.session_label if resolved_session else "Default",
                "entry_mode": entry_mode_val,
                "cgpa": parsed_cgpa,
                "jamb_reg_number": jamb_reg,
                "gender": gender_val,
                "state_of_origin": state_origin,
            })

    return {
        "success": len(errors) == 0,
        "program": {
            "id": str(program.id),
            "name": program.name,
            "code": program.program_code,
            "award_level": program.award_level,
            "duration_years": program.duration_years,
            "department_name": program.department.name,
            "division_name": program.department.division.name,
        },
        "stats": {
            "total_rows": len(raw_rows),
            "valid_count": len(valid_rows),
            "error_count": len(errors),
        },
        "valid_rows": valid_rows,
        "errors": errors,
    }


def commit_student_roster_bulk(
    institution: Institution,
    program_id: str,
    valid_rows: List[Dict[str, Any]],
    default_password_scheme: str = "matric",
) -> Dict[str, Any]:
    """
    Executes an atomic transactional commit of validated student records,
    initializing User accounts and linked StudentProfile instances.
    """
    program = AcademicProgram.objects.get(id=program_id, institution=institution)

    created_users = 0
    created_profiles = 0
    updated_profiles = 0

    with transaction.atomic():
        for r in valid_rows:
            email = r["email"].lower().strip()
            matric = r["matric_number"].strip()
            name = r["name"].strip()

            # 1. Get or create User
            user, u_created = User.objects.get_or_create(
                email=email,
                defaults={"name": name},
            )
            if u_created or not user.has_usable_password():
                initial_pwd = matric.lower() if default_password_scheme == "matric" else "Nexus1234!@#"
                user.set_password(initial_pwd)
                user.name = name
                user.save()
                created_users += 1

            session_id = r.get("entry_session_id")
            session = AcademicSession.objects.filter(id=session_id).first() if session_id else None

            # 2. Update or create StudentProfile
            profile, p_created = StudentProfile.objects.update_or_create(
                user=user,
                defaults={
                    "institution": institution,
                    "program": program,
                    "matric_number": matric,
                    "jamb_reg_number": r.get("jamb_reg_number", ""),
                    "entry_session": session,
                    "entry_mode": r.get("entry_mode", EntryMode.UTME),
                    "year_of_study": r.get("year_of_study", 1),
                    "cgpa": r.get("cgpa"),
                    "phone_number": r.get("phone_number", ""),
                    "gender": r.get("gender", "MALE"),
                    "state_of_origin": r.get("state_of_origin", ""),
                    "is_verified_student": True,
                },
            )

            if p_created:
                created_profiles += 1
            else:
                updated_profiles += 1

    return {
        "success": True,
        "message": f"Successfully ingested {len(valid_rows)} students for {program.name}.",
        "stats": {
            "created_users": created_users,
            "created_profiles": created_profiles,
            "updated_profiles": updated_profiles,
            "total_processed": len(valid_rows),
            "program_name": program.name,
            "department_name": program.department.name,
        },
    }
